import sys
import traceback

try:
    from flask import Flask, render_template, request, redirect, url_for, flash, g, session, send_file
    import sqlite3
    from datetime import date, datetime
    import calendar
    from werkzeug.security import check_password_hash, generate_password_hash
    from functools import wraps
    from flask_mail import Mail, Message
    from itsdangerous import URLSafeTimedSerializer
    from collections import defaultdict
    from dotenv import load_dotenv
    import os
    import io
    import pandas as pd
    from unicodedata import normalize, category
    from flask import jsonify
    from flask_cors import CORS

    # ✅ DETECCIÓN DE RUTAS PARA EJECUTABLE
    if getattr(sys, 'frozen', False):
        # Estamos en el ejecutable
        base_path = sys._MEIPASS
        template_folder = os.path.join(base_path, 'templates')
        static_folder = os.path.join(base_path, 'static')
        PROJECT_ROOT = os.path.dirname(sys.executable)
    else:
        # Estamos en desarrollo
        base_path = os.path.dirname(os.path.abspath(__file__))
        template_folder = os.path.join(base_path, 'templates')
        static_folder = os.path.join(base_path, 'static')
        PROJECT_ROOT = os.path.dirname(base_path)

    app = Flask(__name__, 
                template_folder=template_folder,
                static_folder=static_folder)
    CORS(app)
    app.secret_key = "dev"

    DATABASE = os.path.join(PROJECT_ROOT, "asistencia.db")
    print("Usando base de datos:", DATABASE)

    ALLOWED_SUPERVISORS = ("Directora", "Subdirector Mañana", "Subdirector Tarde", "Ninguno")
    ALLOWED_CONDICIONES = ("Contratado", "Nombrado")
    ALLOWED_TURNOS = ("Mañana", "Tarde", "Completo", "Variado")

    # ✅ CAMBIO CRÍTICO 1: Conexión con WAL y timeout aumentado
    def get_db():
        db = getattr(g, "_db", None)
        if db is None:
            # ✅ timeout=30 evita bloqueos en red
            db = g._db = sqlite3.connect(DATABASE, timeout=30)
            db.row_factory = sqlite3.Row
            
            # ✅ ESTAS LÍNEAS SOLUCIONAN EL PROBLEMA DEL .db-journal
            db.execute("PRAGMA journal_mode=WAL;")
            db.execute("PRAGMA synchronous=NORMAL;")
            db.execute("PRAGMA foreign_keys = ON;")
        return db

    @app.teardown_appcontext
    def close_db(_):
        db = getattr(g, "_db", None)
        if db is not None:
            db.close()

    def current_user():
        uid = session.get("user_id")
        if not uid:
            return None
        db = get_db()
        return db.execute("SELECT ID_USUARIO, NOMBRE, CORREO, ROL FROM USUARIO WHERE ID_USUARIO = ?", (uid,)).fetchone()

    def login_required(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user():
                flash("Debe iniciar sesión para acceder a esta página.", "warning")
                return redirect(url_for("login", next=request.path))
            return f(*args, **kwargs)
        return wrapped

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            correo = request.form.get("correo", "").strip().lower()
            password = request.form.get("password", "")
            db = get_db()
            user = db.execute("SELECT ID_USUARIO, NOMBRE, CORREO, CONTRASEÑA, ROL FROM USUARIO WHERE lower(CORREO) = ?", (correo,)).fetchone()
            if user and check_password_hash(user["CONTRASEÑA"], password):
                session.clear()
                session["user_id"] = user["ID_USUARIO"]
                session["user_name"] = user["NOMBRE"]
                session["user_role"] = user["ROL"]
                flash(f"Bienvenido {user['NOMBRE']}", "success")
                next_url = request.args.get("next")
                if not next_url or next_url == url_for("login") or next_url == url_for("usuarios"):
                    next_url = url_for("menu")
                return redirect(next_url)
            flash("Correo o contraseña incorrectos.", "danger")
            return render_template("login.html", correo=correo)
        return render_template("login.html")

    @app.route("/logout")
    def logout():
        session.clear()
        flash("Sesión cerrada.", "info")
        return redirect(url_for("login"))

    load_dotenv()

    app.config.update(
        MAIL_SERVER=os.getenv("MAIL_SERVER", "smtp.gmail.com"),
        MAIL_PORT=int(os.getenv("MAIL_PORT", 587)),
        MAIL_USE_TLS=os.getenv("MAIL_USE_TLS", "True") == "True",
        MAIL_USERNAME=os.getenv("MAIL_USERNAME"),
        MAIL_PASSWORD=os.getenv("MAIL_PASSWORD"),
        MAIL_DEFAULT_SENDER=(
            os.getenv("MAIL_DEFAULT_SENDER_NAME", "Soporte Asistencias"),
            os.getenv("MAIL_DEFAULT_SENDER_EMAIL")
        ),
        MAIL_SUPPRESS_SEND=False
    )

    mail = Mail(app)

    def generate_token(email):
        s = URLSafeTimedSerializer(app.secret_key)
        return s.dumps(email, salt="reset-password")

    def verify_token(token, expiration=3600):
        s = URLSafeTimedSerializer(app.secret_key)
        try:
            email = s.loads(token, salt="reset-password", max_age=expiration)
        except Exception:
            return None
        return email

    @app.route("/reset_password", methods=["GET", "POST"])
    def reset_password():
        if request.method == "POST":
            correo = request.form.get("correo", "").strip().lower()
            db = get_db()
            user = db.execute(
                "SELECT 1 FROM USUARIO WHERE lower(CORREO) = ?", (correo,)
            ).fetchone()

            if user:
                try:
                    token = generate_token(correo)
                    reset_url = url_for('reset_with_token', token=token, _external=True)

                    msg = Message(
                        subject="Restablecer contraseña",
                        recipients=[correo],
                        charset="utf-8"
                    )
                    msg.html = f"""
                    <div style="text-align:center; font-weight:bold; font-family:Arial, sans-serif; color:#333;">
                    <p style="font-size:16px; margin-bottom:10px;">Hola,</p>
                    <p style="font-size:14px; margin-bottom:10px;">
                        Haz clic en el siguiente botón para restablecer tu contraseña:
                    </p>
                    <p>
                        <a href="{reset_url}" style="
                        background-color:#007bff;
                        color:white;
                        padding:10px 20px;
                        text-decoration:none;
                        border-radius:5px;
                        display:inline-block;
                        font-weight:bold;
                        font-size:16px;
                        ">Restablecer contraseña</a>
                    </p>
                    <p style="font-size:12px; margin-top:20px;">
                        Si no solicitaste este cambio, puedes ignorar este correo.
                    </p>
                    <p style="font-size:12px;">
                        Este enlace expirará en 1 hora.
                    </p>
                    </div>
                    """

                    mail.send(msg)
                    flash("El correo existe, se enviarán instrucciones para restablecer la contraseña.", "info")
                    return redirect(url_for("login"))
                except Exception as e:
                    import traceback
                    print("ERROR enviando correo:", e)
                    traceback.print_exc()
                    flash("Ocurrió un error al enviar el correo. Intenta nuevamente.", "danger")
                    return render_template("reset_password.html", correo=correo)
            else:
                flash("El correo no está registrado en el sistema.", "danger")
                return render_template("reset_password.html", correo=correo)

        return render_template("reset_password.html", correo="")

    @app.route("/reset/<token>", methods=["GET", "POST"])
    def reset_with_token(token):
        email = verify_token(token)
        if not email:
            flash("El enlace para restablecer la contraseña es inválido o ha expirado. Por favor, solicita uno nuevo.", "danger")
            return redirect(url_for("reset_password"))

        if request.method == "POST":
            password = request.form.get("password", "")
            password2 = request.form.get("password2", "")
            if not password or password != password2:
                flash("Las contraseñas no coinciden.", "danger")
                return render_template("reset_with_token.html", token=token)

            db = get_db()
            db.execute(
                "UPDATE USUARIO SET CONTRASEÑA = ? WHERE lower(CORREO) = ?",
                (generate_password_hash(password), email)
            )
            db.commit()
            flash("Contraseña actualizada correctamente. Ahora puedes iniciar sesión.", "success")
            return redirect(url_for("login"))

        return render_template("reset_with_token.html", token=token)

    @app.route("/register", methods=["GET", "POST"])
    def register():
        if request.method == "POST":
            nombre = request.form.get("nombre", "").strip()
            correo = request.form.get("correo", "").strip().lower()
            password = request.form.get("password", "")
            password2 = request.form.get("password2", "")
            rol = request.form.get("rol", "")

            db = get_db()
            if not nombre or not correo or not password or not password2:
                flash("Completa todos los campos.", "warning")
                return render_template("register.html", nombre=nombre, correo=correo)
            if password != password2:
                flash("Las contraseñas no coinciden.", "danger")
                return render_template("register.html", nombre=nombre, correo=correo)
            if db.execute("SELECT 1 FROM USUARIO WHERE lower(CORREO) = ?", (correo,)).fetchone():
                flash("Ya existe una cuenta con ese correo.", "danger")
                return render_template("register.html", nombre=nombre, correo=correo)
            hashpw = generate_password_hash(password)
            db.execute("INSERT INTO USUARIO (NOMBRE, CORREO, CONTRASEÑA, ROL) VALUES (?, ?, ?, ?)",
                    (nombre, correo, hashpw, rol))
            db.commit()
            flash("Cuenta creada correctamente. Ahora puedes iniciar sesión.", "success")
            return redirect(url_for("login"))
        return render_template("register.html")

    @app.route("/menu")
    @login_required
    def menu():
        return render_template("menu.html")

    @app.route("/usuarios")
    @login_required
    def usuarios():
        db = get_db()
        rows = db.execute("""
            SELECT ID_USUARIO, NOMBRE, CORREO, ROL
            FROM USUARIO
            ORDER BY NOMBRE
        """).fetchall()
        return render_template("usuarios_lista.html", rows=rows)
        
    @app.route("/usuarios/eliminar/<int:id_usuario>", methods=["POST"])
    @login_required
    def usuarios_eliminar(id_usuario):
        db = get_db()
        db.execute("DELETE FROM USUARIO WHERE ID_USUARIO = ?", (id_usuario,))
        db.commit()

        if session.get("user_id") == id_usuario:
            session.clear()
            flash("Has eliminado tu usuario. Por favor, inicia sesión nuevamente.", "info")
            return redirect(url_for("login"))

        flash("Usuario eliminado correctamente.", "success")
        return redirect(url_for("usuarios"))

    @app.route("/alumnos")
    @login_required
    def alumnos_lista():
        nivel = request.args.get("nivel", "")
        grado = request.args.get("grado", "")
        seccion = request.args.get("seccion", "")
        dni = request.args.get("dni", "")
        activo = request.args.get("activo", "")
        page = request.args.get("page", 1, type=int)

        db = get_db()

        niveles = db.execute(
            "SELECT DISTINCT NIVEL FROM ESTUDIANTE WHERE NIVEL IS NOT NULL AND NIVEL != '' ORDER BY NIVEL"
        ).fetchall()

        if nivel:
            grados = db.execute(
                "SELECT DISTINCT GRADO FROM ESTUDIANTE WHERE NIVEL = ? AND GRADO IS NOT NULL AND GRADO != '' ORDER BY GRADO",
                (nivel,)
            ).fetchall()
        else:
            grados = db.execute(
                "SELECT DISTINCT GRADO FROM ESTUDIANTE WHERE GRADO IS NOT NULL AND GRADO != '' ORDER BY GRADO"
            ).fetchall()

        if nivel and grado:
            secciones = db.execute(
                "SELECT DISTINCT SECCION FROM ESTUDIANTE WHERE NIVEL = ? AND GRADO = ? AND SECCION IS NOT NULL AND SECCION != '' ORDER BY SECCION",
                (nivel, grado)
            ).fetchall()
        elif grado:
            secciones = db.execute(
                "SELECT DISTINCT SECCION FROM ESTUDIANTE WHERE GRADO = ? AND SECCION IS NOT NULL AND SECCION != '' ORDER BY SECCION",
                (grado,)
            ).fetchall()
        else:
            secciones = db.execute(
                "SELECT DISTINCT SECCION FROM ESTUDIANTE WHERE SECCION IS NOT NULL AND SECCION != '' ORDER BY SECCION"
            ).fetchall()

        has_search = any([nivel, grado, seccion, dni, activo])

        if not has_search:
            alumnos = []
            total_rows = 0
            total_pages = 1
        else:
            conditions = []
            params = []
            if nivel:
                conditions.append("NIVEL = ?"); params.append(nivel)
            if grado:
                conditions.append("GRADO = ?"); params.append(grado)
            if seccion:
                conditions.append("SECCION = ?"); params.append(seccion)
            if dni:
                conditions.append("DNI_ALUMNO LIKE ?"); params.append(f"%{dni}%")
            if activo != "":
                conditions.append("ACTIVO = ?"); params.append(activo)

            where_clause = (" WHERE " + " AND ".join(conditions)) if conditions else ""

            count_sql = "SELECT COUNT(*) as count FROM ESTUDIANTE" + where_clause
            total_rows = db.execute(count_sql, params).fetchone()["count"]

            per_page = 50
            offset = (page - 1) * per_page
            total_pages = (total_rows + per_page - 1) // per_page if total_rows > 0 else 1

            query = (
                "SELECT DNI_ALUMNO AS DNI, NOMBRE, APELLIDO, NIVEL, GRADO, SECCION, ACTIVO "
                "FROM ESTUDIANTE" + where_clause + " ORDER BY APELLIDO, NOMBRE LIMIT ? OFFSET ?"
            )
            alumnos = db.execute(query, params + [per_page, offset]).fetchall()

        return render_template(
            "alumnos_lista.html",
            alumnos=alumnos,
            niveles=niveles,
            grados=grados,
            secciones=secciones,
            nivel_actual=nivel,
            grado_actual=grado,
            seccion_actual=seccion,
            dni_actual=dni,
            activo_actual=activo,
            page=page,
            total_pages=total_pages,
            total_rows=total_rows,
            has_search=has_search
        )

    @app.route("/alumnos/bulk_edit", methods=["GET", "POST"])
    @login_required
    def alumnos_bulk_edit():
        db = get_db()

        def has_updates(form):
            return bool(
                form.get("nivel", "").strip() or
                form.get("grado", "").strip() or
                form.get("seccion", "").strip() or
                (form.get("activo", "") != "")
            )

        if request.method == "POST":
            selected_dnis = request.form.getlist("selected_alumnos")

            if not selected_dnis and session.get("bulk_selected"):
                selected_dnis = session.pop("bulk_selected", [])

            app.logger.debug("BULK_EDIT POST - request.form keys: %s", list(request.form.keys()))
            app.logger.debug("BULK_EDIT POST - selected_dnis: %s", selected_dnis)

            if not selected_dnis:
                flash("No se seleccionaron alumnos para editar.", "warning")
                return redirect(url_for("alumnos_lista"))

            if not has_updates(request.form):
                return render_template("alumnos_bulk_edit.html",
                                    selected_count=len(selected_dnis),
                                    selected_dnis=selected_dnis)

            nuevo_nivel = request.form.get("nivel", "").strip()
            nuevo_grado = request.form.get("grado", "").strip()
            nuevo_seccion = request.form.get("seccion", "").strip()
            nuevo_estado = request.form.get("activo", "")

            app.logger.debug("BULK_EDIT APPLY - cambios solicitados: nivel=%s grado=%s seccion=%s activo=%s",
                            nuevo_nivel, nuevo_grado, nuevo_seccion, nuevo_estado)

            updates = []
            params = []

            if nuevo_nivel:
                updates.append("NIVEL = ?")
                params.append(nuevo_nivel)
            if nuevo_grado:
                updates.append("GRADO = ?")
                params.append(nuevo_grado)
            if nuevo_seccion:
                updates.append("SECCION = ?")
                params.append(nuevo_seccion)
            if nuevo_estado != "":
                try:
                    params.append(int(nuevo_estado))
                    updates.append("ACTIVO = ?")
                except ValueError:
                    flash("Valor inválido para el estado.", "warning")
                    return render_template("alumnos_bulk_edit.html",
                                        selected_count=len(selected_dnis),
                                        selected_dnis=selected_dnis)

            if not updates:
                flash("Debe seleccionar al menos un campo para actualizar.", "warning")
                return render_template("alumnos_bulk_edit.html",
                                    selected_count=len(selected_dnis),
                                    selected_dnis=selected_dnis)

            placeholders = ",".join(["?" for _ in selected_dnis])
            params.extend(selected_dnis)

            sql = f"UPDATE ESTUDIANTE SET {', '.join(updates)} WHERE DNI_ALUMNO IN ({placeholders})"
            app.logger.debug("BULK_EDIT APPLY - SQL: %s - params: %s", sql, params)

            try:
                db.execute(sql, params)
                db.commit()
            except Exception as e:
                app.logger.exception("Error aplicando cambios masivos a alumnos")
                flash("Ocurrió un error al aplicar los cambios. Revisa los logs.", "danger")
                return render_template("alumnos_bulk_edit.html",
                                    selected_count=len(selected_dnis),
                                    selected_dnis=selected_dnis)

            flash(f"Se actualizaron {len(selected_dnis)} alumnos correctamente.", "success")
            return redirect(url_for("alumnos_lista"))

        selected_dnis = request.args.getlist("selected_alumnos") or session.pop("bulk_selected", [])

        if not selected_dnis:
            flash("No se seleccionaron alumnos para editar.", "warning")
            return redirect(url_for("alumnos_lista"))

        return render_template("alumnos_bulk_edit.html",
                            selected_count=len(selected_dnis),
                            selected_dnis=selected_dnis)

    @app.route("/alumnos/nuevo", methods=["GET", "POST"])
    @login_required
    def alumnos_nuevo():
        if request.method == "POST":
            dni = request.form.get("dni", "").strip()
            nombre = request.form.get("nombre", "").strip()
            apellido = request.form.get("apellido", "").strip()
            nivel = request.form.get("nivel", "").strip()
            grado = request.form.get("grado", "").strip()
            seccion = request.form.get("seccion", "").strip()
            activo = request.form.get("activo", "1")

            errores = []
            if not dni:
                errores.append("El DNI es obligatorio.")
            elif len(dni) != 8 or not dni.isdigit():
                errores.append("El DNI debe tener 8 dígitos.")
            if not nombre:
                errores.append("El nombre es obligatorio.")
            if not apellido:
                errores.append("El apellido es obligatorio.")
            if not nivel:
                errores.append("El nivel es obligatorio.")
            if not grado:
                errores.append("El grado es obligatorio.")
            if not seccion:
                errores.append("La sección es obligatoria.")

            if errores:
                for error in errores:
                    flash(error, "warning")
                return render_template("alumnos_form.html", modo="nuevo", alumno=request.form)

            db = get_db()
            existente = db.execute("SELECT 1 FROM ESTUDIANTE WHERE DNI_ALUMNO = ?", (dni,)).fetchone()
            if existente:
                flash("Ya existe un alumno con ese DNI.", "danger")
                return render_template("alumnos_form.html", modo="nuevo", alumno=request.form)

            db.execute("""
                INSERT INTO ESTUDIANTE (DNI_ALUMNO, NOMBRE, APELLIDO, NIVEL, GRADO, SECCION, ACTIVO)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (dni, nombre, apellido, nivel, grado, seccion, int(activo)))
            db.commit()

            flash("Alumno registrado correctamente.", "success")
            return redirect(url_for("alumnos_lista"))

        return render_template("alumnos_form.html", modo="nuevo", alumno=None)

    @app.route("/alumnos/editar/<dni>", methods=["GET", "POST"])
    @login_required
    def alumnos_editar(dni):
        db = get_db()
        if request.method == "POST":
            nombre = request.form.get("nombre", "").strip()
            apellido = request.form.get("apellido", "").strip()
            nivel = request.form.get("nivel", "").strip()
            grado = request.form.get("grado", "").strip()
            seccion = request.form.get("seccion", "").strip()
            activo = request.form.get("activo", "1")

            errores = []
            if not nombre:
                errores.append("El nombre es obligatorio.")
            if not apellido:
                errores.append("El apellido es obligatorio.")
            if not nivel:
                errores.append("El nivel es obligatorio.")
            if not grado:
                errores.append("El grado es obligatorio.")
            if not seccion:
                errores.append("La sección es obligatoria.")

            if errores:
                for error in errores:
                    flash(error, "warning")
                alumno = dict(request.form)
                alumno["DNI"] = dni
                return render_template("alumnos_form.html", modo="editar", alumno=alumno)

            db.execute("""
                UPDATE ESTUDIANTE
                SET NOMBRE = ?, APELLIDO = ?, NIVEL = ?, GRADO = ?, SECCION = ?, ACTIVO = ?
                WHERE DNI_ALUMNO = ?
            """, (nombre, apellido, nivel, grado, seccion, int(activo), dni))
            db.commit()

            flash("Datos del alumno actualizados.", "success")
            return redirect(url_for("alumnos_lista"))

        alumno = db.execute("""
            SELECT DNI_ALUMNO AS DNI, NOMBRE, APELLIDO, NIVEL, GRADO, SECCION, ACTIVO
            FROM ESTUDIANTE
            WHERE DNI_ALUMNO = ?
        """, (dni,)).fetchone()

        if not alumno:
            flash("Alumno no encontrado.", "danger")
            return redirect(url_for("alumnos_lista"))

        return render_template("alumnos_form.html", modo="editar", alumno=alumno)

    @app.route("/alumnos/eliminar/<dni>", methods=["POST"])
    @login_required
    def alumnos_eliminar(dni):
        db = get_db()
        db.execute("DELETE FROM ESTUDIANTE WHERE DNI_ALUMNO = ?", (dni,))
        db.commit()
        flash("Alumno eliminado correctamente.", "success")
        return redirect(url_for("alumnos_lista"))

    @app.route("/personal")
    @login_required
    def personal_lista():
        turno = request.args.get("turno", "").strip()
        q = request.args.get("q", "").strip()
        page = request.args.get("page", 1, type=int)

        db = get_db()

        base_where = " WHERE 1=1"
        params = []

        if turno:
            base_where += " AND TURNO = ?"
            params.append(turno)

        if q:
            like_q = f"%{q}%"
            base_where += " AND (DNI LIKE ? OR NOMBRE_AP_BUSQUEDA LIKE ? OR NOMBRE LIKE ? OR APELLIDO LIKE ?)"
            params.extend([like_q, like_q, like_q, like_q])

        count_sql = "SELECT COUNT(*) as count FROM PERSONAL" + base_where
        total_rows = db.execute(count_sql, params).fetchone()["count"] or 0

        per_page = 50
        offset = (page - 1) * per_page
        total_pages = (total_rows + per_page - 1) // per_page if total_rows > 0 else 1

        query = (
            "SELECT DNI AS DNI, NOMBRE, APELLIDO, CARGO, SUPERVISOR, CONDICION, TURNO, HORAS_JORNADA, NOMBRE_AP_BUSQUEDA "
            "FROM PERSONAL"
            + base_where
            + " ORDER BY APELLIDO, NOMBRE"
            + f" LIMIT {per_page} OFFSET {offset}"
        )

        rows = db.execute(query, params).fetchall()

        return render_template(
            "personal_lista.html",
            rows=rows,
            turno_actual=turno,
            q_actual=q,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            total_rows=total_rows,
        )

    @app.route("/personal/nuevo", methods=["GET", "POST"])
    @login_required
    def personal_nuevo():
        if request.method == "POST":
            dni = request.form.get("dni", "").strip()
            nombre = request.form.get("nombre", "").strip()
            apellido = request.form.get("apellido", "").strip()
            cargo = request.form.get("cargo", "").strip()
            supervisor = request.form.get("supervisor", "").strip() or "Ninguno"
            condicion = request.form.get("condicion", "").strip()
            turno = request.form.get("turno", "").strip()
            horas_jornada = request.form.get("horas_jornada", "").strip()

            if supervisor not in ALLOWED_SUPERVISORS:
                supervisor = "Ninguno"
            if condicion not in ALLOWED_CONDICIONES:
                condicion = ALLOWED_CONDICIONES[0]
            if turno not in ALLOWED_TURNOS:
                turno = ""

            errores = []
            if not dni or len(dni) != 8 or not dni.isdigit():
                errores.append("DNI debe tener 8 dígitos numéricos.")
            if not nombre:
                errores.append("El nombre es obligatorio.")
            if not apellido:
                errores.append("El apellido es obligatorio.")

            if errores:
                for e in errores:
                    flash(e, "warning")
                return render_template("personal_form.html", modo="nuevo", persona=request.form, horarios=None)

            db = get_db()
            existente = db.execute("SELECT 1 FROM PERSONAL WHERE DNI = ?", (dni,)).fetchone()
            if existente:
                flash("Ya existe un registro con ese DNI.", "danger")
                return render_template("personal_form.html", modo="nuevo", persona=request.form, horarios=None)

            texto_busq = quitar_tildes(f"{nombre} {apellido}".lower())

            db.execute("""
                INSERT INTO PERSONAL
                    (DNI, NOMBRE, APELLIDO, CARGO, SUPERVISOR, CONDICION, TURNO, HORAS_JORNADA, NOMBRE_AP_BUSQUEDA)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (dni, nombre, apellido, cargo, supervisor, condicion, turno, horas_jornada, texto_busq))

            lunes = request.form.get("lunes", "").strip() or None
            martes = request.form.get("martes", "").strip() or None
            miercoles = request.form.get("miercoles", "").strip() or None
            jueves = request.form.get("jueves", "").strip() or None
            viernes = request.form.get("viernes", "").strip() or None
            sabado = request.form.get("sabado", "").strip() or None
            domingo = request.form.get("domingo", "").strip() or None

            db.execute("""
                INSERT INTO HORARIO_INGRESO
                    (DNI, LUNES, MARTES, MIERCOLES, JUEVES, VIERNES, SABADO, DOMINGO)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (dni, lunes, martes, miercoles, jueves, viernes, sabado, domingo))

            db.commit()
            flash("Personal y horarios registrados correctamente.", "success")
            return redirect(url_for("personal_lista"))

        return render_template("personal_form.html", modo="nuevo", persona=None, horarios=None)

    @app.route("/personal/editar/<dni>", methods=["GET", "POST"])
    @login_required
    def personal_editar(dni):
        db = get_db()
        if request.method == "POST":
            nombre = request.form.get("nombre", "").strip()
            apellido = request.form.get("apellido", "").strip()
            cargo = request.form.get("cargo", "").strip()
            supervisor = request.form.get("supervisor", "").strip() or "Ninguno"
            condicion = request.form.get("condicion", "").strip()
            turno = request.form.get("turno", "").strip()
            horas_jornada = request.form.get("horas_jornada", "").strip()

            if supervisor not in ALLOWED_SUPERVISORS:
                supervisor = "Ninguno"
            if condicion not in ALLOWED_CONDICIONES:
                condicion = ALLOWED_CONDICIONES[0]
            if turno not in ALLOWED_TURNOS:
                turno = ""

            errores = []
            if not nombre:
                errores.append("El nombre es obligatorio.")
            if not apellido:
                errores.append("El apellido es obligatorio.")

            if errores:
                for e in errores:
                    flash(e, "warning")
                persona = dict(request.form)
                persona["DNI"] = dni
                horarios_row = db.execute("SELECT * FROM HORARIO_INGRESO WHERE DNI = ?", (dni,)).fetchone()
                horarios = dict(horarios_row) if horarios_row else {}
                return render_template("personal_form.html", modo="editar", persona=persona, horarios=horarios)

            texto_busq = quitar_tildes(f"{nombre} {apellido}".lower())

            db.execute("""
                UPDATE PERSONAL
                SET NOMBRE = ?, APELLIDO = ?, CARGO = ?, SUPERVISOR = ?,
                    CONDICION = ?, TURNO = ?, HORAS_JORNADA = ?, NOMBRE_AP_BUSQUEDA = ?
                WHERE DNI = ?
            """, (nombre, apellido, cargo, supervisor, condicion, turno, horas_jornada, texto_busq, dni))

            lunes = request.form.get("lunes", "").strip() or None
            martes = request.form.get("martes", "").strip() or None
            miercoles = request.form.get("miercoles", "").strip() or None
            jueves = request.form.get("jueves", "").strip() or None
            viernes = request.form.get("viernes", "").strip() or None
            sabado = request.form.get("sabado", "").strip() or None
            domingo = request.form.get("domingo", "").strip() or None

            horario_existe = db.execute("SELECT 1 FROM HORARIO_INGRESO WHERE DNI = ?", (dni,)).fetchone()

            if horario_existe:
                db.execute("""
                    UPDATE HORARIO_INGRESO
                    SET LUNES = ?, MARTES = ?, MIERCOLES = ?, JUEVES = ?, 
                        VIERNES = ?, SABADO = ?, DOMINGO = ?
                    WHERE DNI = ?
                """, (lunes, martes, miercoles, jueves, viernes, sabado, domingo, dni))
            else:
                db.execute("""
                    INSERT INTO HORARIO_INGRESO
                        (DNI, LUNES, MARTES, MIERCOLES, JUEVES, VIERNES, SABADO, DOMINGO)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (dni, lunes, martes, miercoles, jueves, viernes, sabado, domingo))

            db.commit()
            flash("Datos de personal y horarios actualizados.", "success")
            return redirect(url_for("personal_lista"))

        persona = db.execute("""
            SELECT DNI, NOMBRE, APELLIDO, CARGO, SUPERVISOR, CONDICION, TURNO, HORAS_JORNADA
            FROM PERSONAL
            WHERE DNI = ?
        """, (dni,)).fetchone()
        
        if not persona:
            flash("Personal no encontrado.", "danger")
            return redirect(url_for("personal_lista"))

        horarios_row = db.execute("""
            SELECT LUNES, MARTES, MIERCOLES, JUEVES, VIERNES, SABADO, DOMINGO
            FROM HORARIO_INGRESO
            WHERE DNI = ?
        """, (dni,)).fetchone()

        horarios = dict(horarios_row) if horarios_row else {}

        return render_template("personal_form.html", modo="editar", persona=persona, horarios=horarios)

    @app.route("/personal/eliminar/<dni>", methods=["POST"])
    @login_required
    def personal_eliminar(dni):
        db = get_db()
        db.execute("DELETE FROM PERSONAL WHERE DNI = ?", (dni,))
        db.commit()
        flash("Registro de personal eliminado.", "info")
        return redirect(url_for("personal_lista"))

    def get_days_of_month(year: int, month: int):
        days = []
        cal = calendar.Calendar()
        for day in cal.itermonthdates(year, month):
            if day.month == month:
                days.append(day)
        return days

    def quitar_tildes(s: str) -> str:
        if not s:
            return ""
        return "".join(
            c for c in normalize("NFD", s)
            if category(c) != "Mn"
        )

    @app.route("/reportes/personal", methods=["GET", "POST"])
    @login_required
    def reportes_personal():
        mes_raw = request.args.get("mes", "")
        supervisor = request.args.get("supervisor", "")
        busqueda_raw = request.args.get("busqueda", "") or ""

        has_search = bool(mes_raw or supervisor or busqueda_raw)

        db = get_db()

        if not has_search:
            hoy = date.today()
            mes_str = f"{hoy.year}-{hoy.month:02d}"
            dias = []
            personal_list = []
            asist_dict = defaultdict(dict)
            return render_template(
                "reportes_personal.html",
                personal_list=personal_list,
                dias=dias,
                asist_dict=asist_dict,
                mes_str=mes_str,
                supervisor=supervisor,
                busqueda=busqueda_raw,
                has_search=has_search
            )

        busqueda = quitar_tildes(busqueda_raw.strip().lower())

        if not mes_raw:
            hoy = date.today()
            mes_str = f"{hoy.year}-{hoy.month:02d}"
        else:
            mes_str = mes_raw

        try:
            year, month = map(int, mes_str.split("-"))
            dias = get_days_of_month(year, month)
        except Exception:
            hoy = date.today()
            mes_str = f"{hoy.year}-{hoy.month:02d}"
            dias = []

        cols = [c["name"] for c in db.execute("PRAGMA table_info('PERSONAL')").fetchall()]
        tiene_col_busqueda = "NOMBRE_AP_BUSQUEDA" in cols

        if tiene_col_busqueda:
            query = "SELECT DNI, NOMBRE, APELLIDO, SUPERVISOR, NOMBRE_AP_BUSQUEDA FROM PERSONAL WHERE 1=1"
        else:
            query = "SELECT DNI, NOMBRE, APELLIDO, SUPERVISOR FROM PERSONAL WHERE 1=1"
        params = []

        if supervisor:
            query += " AND SUPERVISOR = ?"
            params.append(supervisor)

        if busqueda:
            if tiene_col_busqueda:
                like_busq = f"%{busqueda}%"
                query += " AND (NOMBRE_AP_BUSQUEDA LIKE ? OR DNI LIKE ?)"
                params.extend([like_busq, like_busq])
            else:
                like_busq_raw = f"%{busqueda_raw}%"
                query += " AND ((NOMBRE || ' ' || APELLIDO) LIKE ? COLLATE NOCASE OR DNI LIKE ?)"
                params.extend([like_busq_raw, like_busq_raw])

        personal_list = db.execute(query, params).fetchall()

        dnis = [p["DNI"] for p in personal_list]
        if not dnis:
            asistencias = []
        else:
            placeholders = ",".join("?" * len(dnis))
            query_asist = f"""
                SELECT DNI, FECHA, HORARIO_ENTRADA, HORA_LLEGADA, HORA_SALIDA
                FROM ASISTENCIA_PERSONAL
                WHERE DNI IN ({placeholders})
                AND substr(FECHA, 1, 7) = ?
            """
            params_asist = dnis + [mes_str]
            asistencias = db.execute(query_asist, params_asist).fetchall()

        asist_dict = defaultdict(dict)
        for a in asistencias:
            asist_dict[a["DNI"]][a["FECHA"]] = (
                a["HORARIO_ENTRADA"],
                a["HORA_LLEGADA"],
                a["HORA_SALIDA"]  # puede ser None
            )

        return render_template(
            "reportes_personal.html",
            personal_list=personal_list,
            dias=dias,
            asist_dict=asist_dict,
            mes_str=mes_str,
            supervisor=supervisor,
            busqueda=busqueda_raw,
            has_search=has_search
        )
    @app.route("/exportar_excel_personal")
    @login_required
    def exportar_excel_personal():
        mes_str = request.args.get("mes", None)
        supervisor = request.args.get("supervisor", None)
        busqueda = request.args.get("busqueda", "").strip().lower()

        if not mes_str:
            hoy = date.today()
            mes_str = f"{hoy.year}-{hoy.month:02d}"

        year, month = map(int, mes_str.split("-"))
        dias = get_days_of_month(year, month)

        db = get_db()

        query = "SELECT DNI, NOMBRE, APELLIDO, SUPERVISOR FROM PERSONAL WHERE 1=1"
        params = []

        if supervisor:
            query += " AND SUPERVISOR = ?"
            params.append(supervisor)

        if busqueda:
            query += " AND (lower(NOMBRE) LIKE ? OR lower(APELLIDO) LIKE ? OR DNI LIKE ?)"
            like_busq = f"%{busqueda}%"
            params.extend([like_busq, like_busq, like_busq])

        personal_list = db.execute(query, params).fetchall()

        dnis = [p["DNI"] for p in personal_list]
        if not dnis:
            asistencias = []
        else:
            placeholders = ",".join("?" * len(dnis))
            query_asist = f"""
                SELECT DNI, FECHA, HORARIO_ENTRADA, HORA_LLEGADA, HORA_SALIDA
                FROM ASISTENCIA_PERSONAL
                WHERE DNI IN ({placeholders})
                AND substr(FECHA, 1, 7) = ?
            """
            params_asist = dnis + [mes_str]
            asistencias = db.execute(query_asist, params_asist).fetchall()

        asist_dict = defaultdict(dict)
        for a in asistencias:
            asist_dict[a["DNI"]][a["FECHA"]] = (
                a["HORARIO_ENTRADA"],
                a["HORA_LLEGADA"],
                a["HORA_SALIDA"]
            )

        # Construir filas
        filas = []
        for idx, p in enumerate(personal_list, start=1):
            fila = {
                "N°": idx,
                "Nombre Completo": f"{p['NOMBRE']} {p['APELLIDO']}"
            }
            for dia in dias:
                fecha_str = dia.strftime("%Y-%m-%d")
                col_name = dia.strftime("%d/%m")

                if p["DNI"] in asist_dict and fecha_str in asist_dict[p["DNI"]]:
                    entrada, llegada, salida = asist_dict[p["DNI"]][fecha_str]

                    # "Iconos" como texto + saltos de línea
                    texto = f"🕐 {entrada}\n✓ {llegada}"
                    if salida:
                        texto += f"\n→ {salida}"
                    fila[col_name] = texto
                else:
                    fila[col_name] = ""
            filas.append(fila)

        df = pd.DataFrame(filas)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Asistencias")

            # --- Estilos con openpyxl ---
            ws = writer.sheets["Asistencias"]

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            # Encabezado
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajustes de columnas
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 30
            for c in range(3, 3 + len(dias)):
                ws.column_dimensions[get_column_letter(c)].width = 16

            # Celdas de asistencia (fondo suave + wrap)
            fill_asistencia = PatternFill("solid", fgColor="E8F4F8")  # celeste muy suave
            align_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

            max_row = ws.max_row
            max_col = ws.max_column

            for r in range(2, max_row + 1):
                ws.row_dimensions[r].height = 45  # para que entren 2-3 líneas
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if c >= 3 and cell.value:     # solo celdas de días con contenido
                        cell.fill = fill_asistencia
                        cell.alignment = align_wrap

            # Leyenda al final (con color)
            start_legend_row = max_row + 2
            ws.cell(row=start_legend_row, column=1, value="LEYENDA").font = Font(bold=True)

            ws.cell(row=start_legend_row + 1, column=1, value="🕐")
            ws.cell(row=start_legend_row + 1, column=2, value="Horario Programado").font = Font(bold=True, color="0563C1")

            ws.cell(row=start_legend_row + 2, column=1, value="✓")
            ws.cell(row=start_legend_row + 2, column=2, value="Hora de Llegada").font = Font(bold=True, color="00B050")

            ws.cell(row=start_legend_row + 3, column=1, value="→")
            ws.cell(row=start_legend_row + 3, column=2, value="Hora de Salida").font = Font(bold=True, color="C00000")

        output.seek(0)

        return send_file(
            output,
            download_name=f"reporte_asistencias_personal_{mes_str}.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def get_weekdays_of_month(year: int, month: int):
        days = []
        cal = calendar.Calendar()
        for day in cal.itermonthdates(year, month):
            if day.month == month and day.weekday() < 5:
                days.append(day)
        return days
    @app.route("/reportes/alumnos", methods=["GET"])
    @login_required
    def reportes_alumnos():
        mes_raw = request.args.get("mes", "")
        nivel = request.args.get("nivel", "")
        grado = request.args.get("grado", "")
        seccion = request.args.get("seccion", "")

        has_search = bool(mes_raw or nivel or grado or seccion)

        db = get_db()

        niveles = db.execute("""
            SELECT DISTINCT NIVEL
            FROM ESTUDIANTE
            WHERE NIVEL IS NOT NULL AND NIVEL != ''
            ORDER BY NIVEL
        """).fetchall()

        if nivel:
            grados = db.execute("""
                SELECT DISTINCT GRADO
                FROM ESTUDIANTE
                WHERE NIVEL = ? AND GRADO IS NOT NULL AND GRADO != ''
                ORDER BY GRADO
            """, (nivel,)).fetchall()
        else:
            grados = db.execute("""
                SELECT DISTINCT GRADO
                FROM ESTUDIANTE
                WHERE GRADO IS NOT NULL AND GRADO != ''
                ORDER BY GRADO
            """).fetchall()

        if nivel and grado:
            secciones = db.execute("""
                SELECT DISTINCT SECCION
                FROM ESTUDIANTE
                WHERE NIVEL = ? AND GRADO = ? AND SECCION IS NOT NULL AND SECCION != ''
                ORDER BY SECCION
            """, (nivel, grado)).fetchall()
        elif grado:
            secciones = db.execute("""
                SELECT DISTINCT SECCION
                FROM ESTUDIANTE
                WHERE GRADO = ? AND SECCION IS NOT NULL AND SECCION != ''
                ORDER BY SECCION
            """, (grado,)).fetchall()
        else:
            secciones = db.execute("""
                SELECT DISTINCT SECCION
                FROM ESTUDIANTE
                WHERE SECCION IS NOT NULL AND SECCION != ''
                ORDER BY SECCION
            """).fetchall()

        if not has_search:
            hoy = date.today()
            mes_str = f"{hoy.year}-{hoy.month:02d}"
            dias = []
            alumnos = []
            asist_dict = {}
            return render_template(
                "reportes_alumnos.html",
                alumnos=alumnos,
                dias=dias,
                asist_dict=asist_dict,
                mes_str=mes_str,
                nivel=nivel,
                grado=grado,
                seccion=seccion,
                niveles=niveles,
                grados=grados,
                secciones=secciones,
                has_search=has_search
            )

        if not mes_raw:
            hoy = date.today()
            mes_str = f"{hoy.year}-{hoy.month:02d}"
        else:
            mes_str = mes_raw

        try:
            year, month = map(int, mes_str.split("-"))
            dias = get_weekdays_of_month(year, month)
        except Exception:
            hoy = date.today()
            mes_str = f"{hoy.year}-{hoy.month:02d}"
            dias = []

        conditions = []
        params = []
        if nivel:
            conditions.append("NIVEL = ?"); params.append(nivel)
        if grado:
            conditions.append("GRADO = ?"); params.append(grado)
        if seccion:
            conditions.append("SECCION = ?"); params.append(seccion)

        where_clause = (" WHERE " + " AND ".join(conditions)) if conditions else ""

        alumnos_sql = f"""
            SELECT DNI_ALUMNO, NOMBRE, APELLIDO, NIVEL, GRADO, SECCION
            FROM ESTUDIANTE
            {where_clause + (" AND ACTIVO = 1" if where_clause else " WHERE ACTIVO = 1")}
            ORDER BY GRADO, SECCION, APELLIDO, NOMBRE
        """

        alumnos = db.execute(alumnos_sql, params).fetchall()

        dnis = [a["DNI_ALUMNO"] for a in alumnos]
        asist_dict = {}
        if dnis:
            placeholders = ",".join("?" * len(dnis))
            query_asist = f"""
                SELECT DNI_ALUMNO, FECHA, ESTADO
                FROM ASISTENCIA_ALUMNO
                WHERE DNI_ALUMNO IN ({placeholders})
                AND substr(FECHA,1,7) = ?
            """
            params_asist = dnis + [mes_str]
            rows_asist = db.execute(query_asist, params_asist).fetchall()
            for r in rows_asist:
                dni = r["DNI_ALUMNO"]
                fecha = r["FECHA"]
                estado = r["ESTADO"]
                asist_dict.setdefault(dni, {})[fecha] = estado

        return render_template(
            "reportes_alumnos.html",
            alumnos=alumnos,
            dias=dias,
            asist_dict=asist_dict,
            mes_str=mes_str,
            nivel=nivel,
            grado=grado,
            seccion=seccion,
            niveles=niveles,
            grados=grados,
            secciones=secciones,
            has_search=has_search
        )

    @app.route('/exportar_excel_alumnos')
    @login_required
    def exportar_excel_alumnos():
        mes_str = request.args.get('mes', '')
        nivel = request.args.get('nivel', '')
        grado = request.args.get('grado', '')
        seccion = request.args.get('seccion', '')

        if not mes_str:
            hoy = date.today()
            mes_str = f"{hoy.year}-{hoy.month:02d}"

        db = get_db()

        conditions = []
        params = []
        if nivel:
            conditions.append("NIVEL = ?")
            params.append(nivel)
        if grado:
            conditions.append("GRADO = ?")
            params.append(grado)
        if seccion:
            conditions.append("SECCION = ?")
            params.append(seccion)

        where_clause = (" WHERE " + " AND ".join(conditions)) if conditions else ""

        alumnos_sql = f"""
            SELECT DNI_ALUMNO, NOMBRE, APELLIDO
            FROM ESTUDIANTE
            {where_clause + (" AND ACTIVO = 1" if where_clause else " WHERE ACTIVO = 1")}
            ORDER BY GRADO, SECCION, APELLIDO, NOMBRE
        """

        alumnos = db.execute(alumnos_sql, params).fetchall()

        dnis = [a["DNI_ALUMNO"] for a in alumnos]

        try:
            year, month = map(int, mes_str.split("-"))
            num_days = calendar.monthrange(year, month)[1]
            dias_mes = [f"{year}-{month:02d}-{day:02d}" for day in range(1, num_days + 1)]
        except Exception:
            dias_mes = []

        asistencias = defaultdict(dict)
        if dnis:
            placeholders = ",".join(["?"] * len(dnis))
            query_asist = f"""
                SELECT DNI_ALUMNO, FECHA, ESTADO
                FROM ASISTENCIA_ALUMNO
                WHERE DNI_ALUMNO IN ({placeholders})
                AND substr(FECHA,1,7) = ?
            """
            params_asist = dnis + [mes_str]
            rows_asist = db.execute(query_asist, params_asist).fetchall()
            for r in rows_asist:
                dni = r["DNI_ALUMNO"]
                fecha = r["FECHA"]
                estado = r["ESTADO"]
                asistencias[dni][fecha] = estado

        data = []
        for a in alumnos:
            fila = {
                "DNI": a["DNI_ALUMNO"],
                "Nombre": f"{a['NOMBRE']} {a['APELLIDO']}"
            }
            for dia in dias_mes:
                fila[dia] = asistencias.get(a["DNI_ALUMNO"], {}).get(dia, "")
            data.append(fila)

        df = pd.DataFrame(data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Asistencias')
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'asistencias_alumnos_{mes_str}.xlsx'
        )

    @app.route('/api/importar_asistencias_alumnos', methods=['POST'])
    def importar_asistencias_alumnos():
        data = request.get_json()
        if not data or 'asistencias' not in data:
            return jsonify({"ok": False, "error": "JSON inválido"}), 400

        asistencias = data['asistencias']
        dispositivo_id = data.get('dispositivo_id')
        fecha_sync = data.get('fecha_sincronizacion') or datetime.now().isoformat(timespec='seconds')

        conn = get_db()
        cur = conn.cursor()

        insertados = 0
        duplicados = 0
        errores = 0

        for a in asistencias:
            dni = a.get('dni_alumno')
            fecha = a.get('fecha')
            estado = a.get('estado')
            observacion = a.get('observacion')

            if not (dni and fecha and estado):
                errores += 1
                continue

            if estado not in ('A', 'F', 'J'):
                errores += 1
                continue

            cur.execute("""
                SELECT COUNT(*) FROM ASISTENCIA_ALUMNO
                WHERE DNI_ALUMNO = ? AND FECHA = ?
            """, (dni, fecha))
            ya_hay = cur.fetchone()[0]

            if ya_hay:
                duplicados += 1
                continue

            cur.execute("""
                INSERT INTO ASISTENCIA_ALUMNO
                    (DNI_ALUMNO, FECHA, ESTADO, OBSERVACION, SYNCED_AT)
                VALUES (?, ?, ?, ?, ?)
            """, (dni, fecha, estado, observacion, fecha_sync))
            insertados += 1

        conn.commit()

        return jsonify({
            "ok": True,
            "insertados": insertados,
            "duplicados": duplicados,
            "errores": errores
        }), 200

    # ✅ CAMBIO CRÍTICO 2: Ruta de sincronización con conexión directa y timeout
    @app.route('/api/sync/asistencia-alumnos', methods=['POST'])
    def sync_asistencia_alumnos():
        try:
            data = request.get_json(force=True)
            items = data.get('items', [])
            
            if not items:
                return jsonify({"ok": False, "msg": "No se recibieron datos"}), 400

            # ✅ Conexión directa con timeout y WAL
            conn = sqlite3.connect(DATABASE, timeout=30)
            conn.execute("PRAGMA journal_mode=WAL;")
            conn.execute("PRAGMA synchronous=NORMAL;")
            conn.execute("PRAGMA foreign_keys = ON;")
            cur = conn.cursor()
            
            ok_count = 0
            failed = []
            ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            for it in items:
                try:
                    dni = it.get('dni_alumno')
                    fecha = it.get('fecha')
                    estado = it.get('estado', 'A')
                    obs = it.get('observacion', '')

                    cur.execute("""
                        INSERT INTO ASISTENCIA_ALUMNO 
                        (DNI_ALUMNO, FECHA, ESTADO, OBSERVACION, SYNCED_AT)
                        VALUES (?, ?, ?, ?, ?)
                        ON CONFLICT(DNI_ALUMNO, FECHA)
                        DO UPDATE SET
                            ESTADO=excluded.ESTADO,
                            OBSERVACION=excluded.OBSERVACION,
                            SYNCED_AT=excluded.SYNCED_AT
                    """, (dni, fecha, estado, obs, ahora))
                    
                    ok_count += 1
                    print(f"✅ Sincronizado: {dni} - {fecha} - {estado}")
                    
                except Exception as e:
                    print(f"❌ ERROR: {e} | Item: {it}")
                    failed.append({"item": it, "error": str(e)})

            conn.commit()
            conn.close()

            return jsonify({
                "ok": True, 
                "inserted_or_updated": ok_count, 
                "failed": failed
            }), 200

        except Exception as e:
            print(f"❌ ERROR GENERAL: {e}")
            return jsonify({"ok": False, "msg": str(e)}), 500

    @app.route('/api/alumnos', methods=['GET'])
    def api_alumnos():
        conn = get_db()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT DNI_ALUMNO, NOMBRE, APELLIDO, GRADO, SECCION, NIVEL
            FROM ESTUDIANTE
        """)
        rows = [dict(r) for r in cur.fetchall()]
        return jsonify(rows)

    @app.route('/api/grados', methods=['GET'])
    def api_grados():
        conn = get_db()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT GRADO FROM ESTUDIANTE ORDER BY GRADO")
        rows = [dict(r) for r in cur.fetchall()]
        return jsonify(rows)

    @app.route('/api/secciones', methods=['GET'])
    def api_secciones():
        conn = get_db()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT SECCION FROM ESTUDIANTE ORDER BY SECCION")
        rows = [dict(r) for r in cur.fetchall()]
        return jsonify(rows)

    def dia_semana_columna(fecha: datetime) -> str:
        mapping = {
            0: "LUNES", 1: "MARTES", 2: "MIERCOLES",
            3: "JUEVES", 4: "VIERNES", 5: "SABADO", 6: "DOMINGO"
        }
        return mapping[fecha.weekday()]

    @app.route("/ingreso", methods=["GET", "POST"])
    def ingreso_personal():
        msg = None
        detalle = None
        mostrar_confirmacion = False
        dni_confirmacion = None
        
        if request.method == "POST":
            dni = request.form.get("dni", "").strip()
            confirmar_salida = request.form.get("confirmar_salida", "").strip()
            
            if not dni or len(dni) != 8 or not dni.isdigit():
                flash("Ingrese un DNI válido de 8 dígitos.", "warning")
                return redirect(url_for("ingreso_personal"))

            now = datetime.now()
            fecha_str = now.strftime("%Y-%m-%d")
            hora_str = now.strftime("%H:%M")
            col_dia = dia_semana_columna(now)

            db = get_db()

            per = db.execute(
                "SELECT DNI, NOMBRE, APELLIDO FROM PERSONAL WHERE DNI = ?",
                (dni,)
            ).fetchone()
            if not per:
                flash("DNI no encontrado en el personal.", "danger")
                return redirect(url_for("ingreso_personal"))

            registro_hoy = db.execute(
                "SELECT ID_ASIS_PERSONAL, HORA_LLEGADA, HORA_SALIDA, HORARIO_ENTRADA FROM ASISTENCIA_PERSONAL WHERE DNI = ? AND FECHA = ?",
                (dni, fecha_str),
            ).fetchone()

            # CASO 1: Ya entró pero NO tiene salida → mostrar modal o registrar salida
            if registro_hoy and not registro_hoy["HORA_SALIDA"]:
                if confirmar_salida != "si":
                    mostrar_confirmacion = True
                    dni_confirmacion = dni
                    detalle = {
                        "nombre": f"{per['NOMBRE']} {per['APELLIDO']}",
                        "hora_entrada": registro_hoy["HORA_LLEGADA"],  # para el modal
                    }
                    return render_template("ingreso.html",
                                        mostrar_confirmacion=mostrar_confirmacion,
                                        dni_confirmacion=dni_confirmacion,
                                        detalle=detalle)
                else:
                    db.execute(
                        "UPDATE ASISTENCIA_PERSONAL SET HORA_SALIDA = ? WHERE ID_ASIS_PERSONAL = ?",
                        (hora_str, registro_hoy["ID_ASIS_PERSONAL"])
                    )
                    db.commit()

                    msg = "✓ Hora de salida registrada"
                    detalle = {
                        "nombre": f"{per['NOMBRE']} {per['APELLIDO']}",
                        "fecha": fecha_str,
                        # ⭐ hora_llegada para mostrar arriba de la salida en el HTML
                        "hora_llegada": registro_hoy["HORA_LLEGADA"],
                        "hora_salida": hora_str,
                    }
                    flash(msg, "success")
                    return render_template("ingreso.html", msg=msg, detalle=detalle)

            # CASO 2: Ya tiene entrada Y salida
            elif registro_hoy and registro_hoy["HORA_SALIDA"]:
                flash(f"{per['NOMBRE']} {per['APELLIDO']} ya registró entrada y salida hoy.", "info")
                return redirect(url_for("ingreso_personal"))

            # CASO 3: Primera vez del día → registrar entrada
            else:
                row_h = db.execute(
                    f"SELECT {col_dia} AS HORARIO FROM HORARIO_INGRESO WHERE DNI = ?",
                    (dni,)
                ).fetchone()
                horario_prog = row_h["HORARIO"] if row_h and row_h["HORARIO"] else "No definido"

                db.execute(
                    "INSERT INTO ASISTENCIA_PERSONAL (DNI, FECHA, HORARIO_ENTRADA, HORA_LLEGADA) VALUES (?,?,?,?)",
                    (dni, fecha_str, horario_prog, hora_str),
                )
                db.commit()

                msg = "✓ Registro de entrada exitoso"
                detalle = {
                    "nombre": f"{per['NOMBRE']} {per['APELLIDO']}",
                    "fecha": fecha_str,
                    # ⭐ Clave corregida: horario_programado (igual que en el HTML)
                    "horario_programado": horario_prog,
                    "hora_llegada": hora_str,
                }
                flash(msg, "success")

        return render_template("ingreso.html", msg=msg, detalle=detalle)

    @app.get("/")
    def index():
        return redirect(url_for("login"))

    if __name__ == "__main__":
        app.run(debug=False, host='0.0.0.0', port=5000)

except Exception as e:
    print("----------------------------------------------------")
    print("¡ERROR CRÍTICO AL INICIAR LA APLICACIÓN FLASK!")
    print("----------------------------------------------------")
    print(f"Tipo de error: {type(e).__name__}")
    print(f"Mensaje: {e}")
    print("\nTraceback completo:")
    traceback.print_exc(file=sys.stdout)
    print("----------------------------------------------------")
    sys.exit(1)